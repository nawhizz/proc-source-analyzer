import re
import sys
import os
from collections import defaultdict

def analyze_file(file_path, encoding='euc-kr'):
    """
    Pro*C 파일을 분석하여 TB_로 시작하는 테이블과 CRUD 작업을 추출합니다.
    EXEC SQL 블록과 문자열 리터럴(동적 쿼리)을 모두 분석합니다.
    추가로 '프로그램명 : ...' 패턴을 찾아 설명을 추출합니다.
    """
    if not os.path.exists(file_path):
        print(f"Error: File not found - {file_path}")
        return {}, ""

    try:
        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading file: {e}")
        return {}, ""

    table_ops = defaultdict(set)
    source_desc = ""

    # 0. 프로그램명 / 설명 추출
    # 우선순위: 프로그램명 -> 파일명(한글) -> Description
    patterns = [
        r'프로그램\s*명\s*:\s*(.*)',      # 프로그램 명 : ...
        r'파일명\s*\(\s*한글\s*\)\s*:\s*(.*)',  # 파일명(한글) : ...
        r'Description\s*:\s*(.*)',      # Description : ...
        r'Descritpion\s*:\s*(.*)'      # Descritpion : ...
    ]
    
    for pat in patterns:
        match = re.search(pat, content, re.IGNORECASE)
        if match:
            extracted = match.group(1).strip()
            if extracted:
                source_desc = extracted
                break
    
    # 1. EXEC SQL 블록 분석 (정적 쿼리)
    # exec sql 로 시작하고 ; 로 끝나는 블록을 찾음 (줄바꿈 포함)
    exec_sql_pattern = re.compile(r'EXEC\s+SQL\s+(.*?);', re.DOTALL | re.IGNORECASE)

    for match in exec_sql_pattern.finditer(content):
        sql_block = match.group(1)
        extract_table_crud(sql_block, table_ops, source="STATIC")

    # 2. 문자열 리터럴 분석 (동적 쿼리)
    # C언어 스타일의 문자열 연결(String Concatenation)을 처리합니다.
    # 예: "SELECT * " \n " FROM TB_TEST" -> "SELECT *  FROM TB_TEST"
    
    # 패턴 설명:
    # "..." : 첫 번째 문자열 (이스케이프 문자 처리 포함)
    # (?:\s*"...")* : 공백(줄바꿈 포함) 후 이어지는 문자열들이 0개 이상 반복
    concat_string_pattern = re.compile(r'("(?:\\[\s\S]|[^"\\])*"(?:\s*"(?:\\[\s\S]|[^"\\])*")*)')
    
    for match in concat_string_pattern.finditer(content):
        full_match = match.group(1)
        
        # 연결된 문자열들을 하나로 합치기
        # 1. 각 "..." 블록을 찾음
        single_str_pattern = re.compile(r'"((?:\\[\s\S]|[^"\\])*)"')
        parts = single_str_pattern.findall(full_match)
        
        if parts:
            # 2. 하나의 문자열로 결합 (공백 하나로 구분하여 안전하게 연결)
            sql_string = " ".join(parts)
            
            # C-style escape sequence handling (\n, \r, \t -> space)
            # Literal backslash + n/r/t in the source string becomes literal characters in sql_string
            # We replace them with space to allow regex \s+ to match
            sql_string = re.sub(r'\\[nrt]', ' ', sql_string)
            
            # 문자열 안에 TB_, ATA_, EM_ 테이블이 있는지 확인
            if any(prefix in sql_string for prefix in ["TB_", "ATA_", "EM_"]):
                extract_table_crud(sql_string, table_ops, source="DYNAMIC")

    return table_ops, source_desc

def extract_table_crud(sql_text, table_ops, source="UNKNOWN"):
    """
    SQL 텍스트(또는 문자열)에서 TB_, ATA_, EM_ 테이블과 CRUD 키워드를 추출하여 table_ops에 저장합니다.
    단어 유무만 확인하는 것이 아니라, 문맥(INSERT INTO, UPDATE, FROM 등)을 고려하여
    정확한 CRUD 작업을 식별합니다.
    """
    # 대문자로 변환하여 분석
    sql_upper = sql_text.upper()

    # 주석 제거 (힌트 처리 및 오탐 방지)
    # /* ... */ 형태의 주석을 공백으로 교체하여 길이(인덱스) 유지
    # 단순화: -- 주석은 행 단위 처리가 필요하므로 여기서는 /* */ 만 처리 (User Case 대응)
    # 정규식: /* 부터 */ 까지 (Non-greedy)
    sql_clean = re.sub(r'/\*.*?\*/', lambda m: ' ' * len(m.group()), sql_upper, flags=re.DOTALL)
    
    # MERGE 문 특수 처리 (Cleaned SQL 사용)
    if 'MERGE' in sql_clean:
        process_merge_statement(sql_clean, table_ops)
        return

    # 테이블 패턴 (스키마 포함)
    table_pattern = re.compile(r'\b((?:[A-Z0-9_]+\.)?(?:TB_|ATA_|EM_)[A-Z0-9_]+)\b')
    
    # 모든 테이블 등장 위치 찾기 - Clean된 SQL에서 찾음 (주석 내 테이블 무시 효과)
    found_tables = []
    for match in table_pattern.finditer(sql_clean):
        found_tables.append({
            'name': match.group(1),
            'start': match.start(),
            'end': match.end(),
            'ops': set()
        })
    
    if not found_tables:
        return

    # CRUD 타겟 패턴 정의
    # INSERT INTO Table
    insert_pattern = re.compile(r'INSERT\s+INTO\s+((?:[A-Z0-9_]+\.)?(?:TB_|ATA_|EM_)[A-Z0-9_]+)')
    # INSERT Columns Pattern: INSERT INTO Table (...)
    # 괄호 안의 내용을 비글리디 하게 잡되, 줄바꿈 포함
    insert_columns_pattern = re.compile(r'INSERT\s+INTO\s+[A-Z0-9_.]+\s*\((.*?)\)', re.DOTALL)
    
    # UPDATE Table
    update_pattern = re.compile(r'UPDATE\s+((?:[A-Z0-9_]+\.)?(?:TB_|ATA_|EM_)[A-Z0-9_]+)')
    # DELETE [FROM] Table
    delete_pattern = re.compile(r'DELETE\s+(?:FROM\s+)?((?:[A-Z0-9_]+\.)?(?:TB_|ATA_|EM_)[A-Z0-9_]+)')

    # Column Confusion Check: Remove found_tables that are inside INSERT column list
    valid_tables = []
    
    # 1. Find all exclusion zones (INSERT column lists)
    exclusion_zones = []
    for match in insert_columns_pattern.finditer(sql_clean):
        # group(1) has the content inside parens
        start_idx = match.start(1)
        end_idx = match.end(1)
        exclusion_zones.append((start_idx, end_idx))
        
    # 2. Filter found_tables
    for ft in found_tables:
        is_column = False
        for (ex_start, ex_end) in exclusion_zones:
            # If table is completely inside the exclusion zone
            if ft['start'] >= ex_start and ft['end'] <= ex_end:
                is_column = True
                break
        
        if not is_column:
            valid_tables.append(ft)
    
    found_tables = valid_tables
    if not found_tables:
        return

    # 각 패턴별로 매칭되는 테이블의 위치(span)를 찾아 해당 작업 부여
    # Note: sql_clean을 사용하여 힌트가 공백 처리된 상태이므로 INSERT ... INTO 매칭 성공
    def mark_operations(pattern, op_name):
        for match in pattern.finditer(sql_clean):
            # 매칭된 그룹(테이블명)의 span
            try:
                target_start = match.start(1)
                target_end = match.end(1)
                
                # 발견된 테이블 목록에서 이 위치와 일치하는 항목 찾기
                for ft in found_tables:
                    # 완벽히 겹치는지 확인 (같은 위치인지)
                    if ft['start'] == target_start and ft['end'] == target_end:
                        ft['ops'].add(op_name)
            except IndexError:
                pass # 패턴에 그룹이 없는 경우 무시

    mark_operations(insert_pattern, 'INSERT')
    mark_operations(update_pattern, 'UPDATE')
    mark_operations(delete_pattern, 'DELETE')

    # 기본적으로 모든 테이블은 SELECT로 간주하되,
    # 위에서 INSERT/UPDATE/DELETE로 마킹된 적이 없는 경우에만 SELECT를 추가합니다.
    # 단, INSERT/UPDATE/DELETE 구문의 타겟이 아니면 모두 조회(SELECT) 용도로 봅니다.
    # 예: INSERT INTO T1 SELECT * FROM T1;
    # 첫번째 T1: INSERT 마킹됨
    # 두번째 T1: 마킹 안됨 -> SELECT 추가
    
    # [ROBUST SELECT LOGIC]
    # 단순 Default가 아니라, FROM 또는 JOIN 절에 포함되어 있는지 역방향 탐색으로 확인합니다.
    # 이를 통해 SELECT 절, WHERE 절, UPDATE SET 절 등에 있는 컬럼명(ATA_ID 등)이 오탐지되는 것을 방지합니다.
    
    def is_select_source(table_info):
        # 테이블 시작 위치에서 역방향으로 스캔하여, 문장의 구조적 위치를 파악
        start_pos = table_info['start']
        # 최대 1000자 정도 역탐색 (성능 고려)
        lookback_limit = max(0, start_pos - 1000)
        chunk = sql_clean[lookback_limit:start_pos]
        
        # 역순으로 하나씩 토큰을 확인
        # 의미있는 토큰: FROM, JOIN, ',', INSERT, UPDATE, DELETE, SET, WHERE, SELECT, '(', ')'
        # 콤마(',')는 계속 이전 항목을 찾도록 연결해줌 (FROM T1, T2)
        # 하지만 SELECT T1, T2 FROM ... 과 구분하기 위해
        # 콤마를 만나면 계속 뒤로 가되, 궁극적으로 FROM/JOIN을 만나야 함.
        # SELECT, SET, WHERE 등을 만나면 False.
        
        # 정규식으로 토큰화하여 마지막(가장 가까운) 토큰부터 확인
        # \b 키워드 \b 또는 구두점
        tokens = list(re.finditer(r'\b(FROM|JOIN|UPDATE|INSERT|DELETE|SELECT|SET|WHERE|GROUP|ORDER|HAVING|VALUES)\b|[,()]', chunk, re.IGNORECASE))
        
        if not tokens:
            return False
            
        # 뒤에서부터 순회
        paren_depth = 0
        for i in range(len(tokens)-1, -1, -1):
            tok = tokens[i].group().upper()
            
            if tok == ')':
                paren_depth += 1
            elif tok == '(':
                # FROM (Subquery) -> Subquery 닫힘? 아니면 여는 괄호?
                # 역순이므로 ')' 가 먼저 나오고 '(' 가 나중에 나옴.
                # 즉 ')' 는 닫는 괄호(원래 문장 순서상), '(' 는 여는 괄호.
                if paren_depth > 0:
                    paren_depth -= 1
                else:
                    # 괄호 밖으로 나감? 혹은 단순 그룹핑?
                    # FROM (TB) -> OK.
                    # IN (COL) -> OK.
                    # 일단 괄호는 구조적 장벽으로 보고, depth 0일때 괄호에 막히면 중단할 수도 있음.
                    # 하지만 Table 명이 바로 뒤에 나오는 경우 (FROM T1) 에는 괄호가 없음.
                    # (T1) -> T1 앞에 (.
                    pass
            
            if paren_depth > 0:
                continue

            if tok in (',',):
                continue
            
            if tok in ('FROM', 'JOIN'):
                # DELETE FROM 인지 확인해야 함
                # FROM 바로 앞 토큰(i-1)이 DELETE인지 확인
                if i > 0:
                    prev = tokens[i-1].group().upper()
                    if prev == 'DELETE':
                        return False # DELETE target, handled elsewhere
                return True
                
            if tok in ('UPDATE', 'INSERT', 'DELETE', 'SELECT', 'SET', 'WHERE', 'GROUP', 'ORDER', 'HAVING', 'VALUES'):
                # 다른 절의 시작을 만나면 소스 테이블이 아님
                return False
                
        return False

    for ft in found_tables:
        # NHPT. 스키마 제거
        clean_name = ft['name']
        if clean_name.startswith("NHPT."):
            clean_name = clean_name.replace("NHPT.", "")
            
        if not ft['ops']:
            # 명시적 CRUD 타겟이 아닌 경우
            # SELECT Source 인지 문맥 체크
            if is_select_source(ft):
                table_ops[clean_name].add('SELECT')
            
            # 여기서 매칭되지 않으면 (예: SELECT 절의 컬럼, WHERE 절의 컬럼 등) 아무 작업도 부여되지 않음 -> 무시됨.
            
        else:
            # 이미 CRUD 작업이 식별된 경우
            for op in ft['ops']:
                table_ops[clean_name].add(op)

def process_merge_statement(sql_upper, table_ops):
    """
    MERGE 문을 상세 분석하여 Target 테이블에는 INSERT/UPDATE를,
    Source 테이블 등에는 SELECT를 부여합니다.
    """
    # Target Table 추출: MERGE INTO [Table]
    # Target Table 추출: MERGE INTO [Table]
    target_pattern = re.compile(r'MERGE\s+INTO\s+((?:[A-Z0-9_]+\.)?(?:TB_|ATA_|EM_)[A-Z0-9_]+)', re.DOTALL)
    target_match = target_pattern.search(sql_upper)
    
    target_table = None
    if target_match:
        target_table = target_match.group(1)
        # NHPT. 스키마 제거
        if target_table.startswith("NHPT."):
            target_table = target_table.replace("NHPT.", "")
        
        # Target Table Operations
        # 줄바꿈 등이 섞여 있을 수 있으므로 re.DOTALL 사용
        # WHEN MATCHED THEN UPDATE
        if re.search(r'WHEN\s+MATCHED\s+THEN\s+UPDATE', sql_upper, re.DOTALL):
            table_ops[target_table].add('UPDATE')
        
        # WHEN NOT MATCHED THEN INSERT
        if re.search(r'WHEN\s+NOT\s+MATCHED\s+THEN\s+INSERT', sql_upper, re.DOTALL):
            table_ops[target_table].add('INSERT')
            
    # 나머지 테이블 추출 (Source Tables) -> SELECT 취급
    # 전체 테이블 찾기 (리스트로 반환하여 개수 확인)
    all_table_pattern = re.compile(r'\b((?:[A-Z0-9_]+\.)?(?:TB_|ATA_|EM_)[A-Z0-9_]+)\b')
    all_tables_list = all_table_pattern.findall(sql_upper)
    all_tables = set(all_tables_list)
    
    # Target Table 제외 로직 개선
    # Target Table이 SQL 문 내에서 여러 번 등장하면 Source(SELECT)로도 사용된 것으로 간주
    if target_match:
        raw_target_table = target_match.group(1)
        
        # Target Table이 목록에 있고, 등장 횟수가 1번뿐이라면 (Target으로만 사용됨) -> SELECT 목록에서 제외
        # 만약 2번 이상 등장하면 (Target + Source) -> SELECT 목록에 유지
        if raw_target_table in all_tables:
            if all_tables_list.count(raw_target_table) == 1:
                all_tables.remove(raw_target_table)
        
    # 나머지는 모두 SELECT (USING 구문 등)
    for table in all_tables:
        # NHPT. 스키마 제거
        if table.startswith("NHPT."):
            table = table.replace("NHPT.", "")
        table_ops[table].add('SELECT')

import argparse
import glob
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def main():
    parser = argparse.ArgumentParser(description="Pro*C Source Analyzer")
    parser.add_argument("-f", "--file", help="Path to a single Pro*C file to analyze")
    parser.add_argument("-d", "--folder", help="Directory path to scan for *.pc files")
    parser.add_argument("-e", "--excel", help="Output Excel filename (e.g., result.xlsx)")
    parser.add_argument("-m", "--merge", action="store_true", help="Merge cells for same Source Name and Source Desc. in Excel")
    parser.add_argument("-c", "--encoding", default="euc-kr", help="File encoding (default: euc-kr)")
    
    args = parser.parse_args()
    
    if args.excel and not OPENPYXL_AVAILABLE:
        print("Error: 'openpyxl' library is not installed. Please install it using 'pip install openpyxl' or 'uv add openpyxl' to use Excel export.")
        sys.exit(1)

    files_to_process = []
    
    if args.file:
        files_to_process.append(args.file)
    elif args.folder:
        if os.path.exists(args.folder):
            # Recursive search using ** (requires Python 3.5+)
            search_pattern = os.path.join(args.folder, "**", "*.pc")
            files_to_process = glob.glob(search_pattern, recursive=True)
            if not files_to_process:
                print(f"No *.pc files found in: {args.folder} (recursive scan)")
        else:
            print(f"Error: Directory not found - {args.folder}")
            sys.exit(1)
    else:
        parser.print_help()
        sys.exit(1)

    # Collect all results
    all_results = [] # List of tuples: (filename, source_desc, table, operations)

    for file_path in files_to_process:
        result, source_desc = analyze_file(file_path, encoding=args.encoding)
        file_name = os.path.basename(file_path)

        # Console Output
        print(f"\nAnalysis Report for: {file_path}")
        print(f"Source Desc: {source_desc}")
        print(f"{'Table Name':<30} | {'CRUD Operations'}")
        print("-" * 60)
        
        if not result:
            print("No tables found or file error.")
        else:
            sorted_tables = sorted(result.keys())
            for table in sorted_tables:
                ops = ", ".join(sorted(result[table]))
                print(f"{table:<30} | {ops}")
                # Add to results for Excel
                all_results.append((file_name, source_desc, table, ops))
        print("\n" + "="*60)

    # Excel Export
    if args.excel:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Result"
            
            # Header
            from openpyxl.styles import Alignment
            ws.append(["Source Name", "Source Desc.", "Table Name", "CRUD Operations"])
            
            # Data
            # Sort mainly for merging logic (though processing order might be enough, safety first)
            # But here we process file by file, so it's naturally sorted by file processing order.
            
            start_row = 2 # Data starts from row 2
            for row in all_results:
                ws.append(row)
            
            last_row = ws.max_row

            # Merging Logic
            if args.merge and last_row >= start_row:
                # Iterate to find ranges to merge
                # We need to merge Column 1 (Source Name) and Column 2 (Source Desc.)
                
                # Helper to merge a specific column
                def merge_column(col_idx):
                    current_val = ws.cell(row=start_row, column=col_idx).value
                    merge_start = start_row
                    
                    for r in range(start_row + 1, last_row + 2): # Go one past end to handle last block
                        val = ws.cell(row=r, column=col_idx).value if r <= last_row else None
                        
                        if val != current_val:
                            # End of a block
                            if r - 1 > merge_start:
                                ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=r-1, end_column=col_idx)
                                # Center alignment for merged cells
                                cell = ws.cell(row=merge_start, column=col_idx)
                                cell.alignment = Alignment(vertical='center', horizontal='center')
                            
                            current_val = val
                            merge_start = r

                merge_column(1) # Source Name
                merge_column(2) # Source Desc.

            # Auto-adjust column width (simple approximation)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(args.excel)
            print(f"\nExcel file saved successfully to: {args.excel}")
        except Exception as e:
            print(f"\nError saving Excel file: {e}")

if __name__ == "__main__":
    main()
